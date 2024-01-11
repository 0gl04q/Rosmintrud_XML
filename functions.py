import lxml.etree as ET
from datetime import datetime
from openpyxl import load_workbook

# Наименования ПО
DI = {
    "1": "Оказание первой помощи пострадавшим",
    "2": "Использование (применение) средств индивидуальной защиты",
    "3": "Общие вопросы охраны труда и функционирования системы управления охраной труда",
    "4": "Безопасные методы и приемы выполнения работ при воздействии вредных и (или) опасных производственных факторов, источников опасности, идентифицированных в рамках специальной оценки условий труда и оценки профессиональных рисков",
    "6": "Безопасные методы и приемы выполнения земляных работ",
    "7": "Безопасные методы и приемы выполнения ремонтных, монтажных и демонтажных работ зданий и сооружений",
    "8": "Безопасные методы и приемы выполнения работ при размещении, монтаже, техническом обслуживании и ремонте технологического оборудования (включая технологическое оборудование)",
    "9": "Безопасные методы и приемы выполнения работ на высоте",
    "10": "Безопасные методы и приемы выполнения пожароопасных работ",
    "11": "Безопасные методы и приемы выполнения работ в ограниченных и замкнутых пространствах (ОЗП)",
    "12": "Безопасные методы и приемы выполнения строительных работ, в том числе: - окрасочные работы - электросварочные и газосварочные работы",
    "13": "Безопасные методы и приемы выполнения работ, связанных с опасностью воздействия сильнодействующих и ядовитых веществ",
    "14": "Безопасные методы и приемы выполнения газоопасных работ",
    "15": "Безопасные методы и приемы выполнения огневых работ",
    "16": "Безопасные методы и приемы выполнения работ, связанные с эксплуатацией подъемных сооружений",
    "17": "Безопасные методы и приемы выполнения работ, связанные с эксплуатацией тепловых энергоустановок",
    "18": "Безопасные методы и приемы выполнения работ в электроустановках",
    "19": "Безопасные методы и приемы выполнения работ, связанные с эксплуатацией сосудов, работающих под избыточным давлением",
    "20": "Безопасные методы и приемы обращения с животными",
    "21": "Безопасные методы и приемы при выполнении водолазных работ",
    "22": "Безопасные методы и приемы работ по поиску, идентификации, обезвреживанию и уничтожению взрывоопасных предметов",
    "23": "Безопасные методы и приемы работ в непосредственной близости от полотна или проезжей части эксплуатируемых автомобильных и железных дорог",
    "24": "Безопасные методы и приемы работ, на участках с патогенным заражением почвы",
    "25": "Безопасные методы и приемы работ по валке леса в особо опасных условиях",
    "26": "Безопасные методы и приемы работ по перемещению тяжеловесных и крупногабаритных грузов при отсутствии машин соответствующей грузоподъемности и разборке покосившихся и опасных (неправильно уложенных) штабелей круглых лесоматериалов",
    "27": "Безопасные методы и приемы работ с радиоактивными веществами и источниками ионизирующих излучений",
    "28": "Безопасные методы и приемы работ с ручным инструментом, в том числе с пиротехническим",
    "29": "Безопасные методы и приемы работ в театрах"
}

# Путь к файлу
PATH_TO_MAIN = r'\\192.168.10.10\учебный центр\Учеба\Журналы\Журнал регистрации удостоверений АТМ ДОТ 2024.xlsx'
PATH_TO_XML = r'\\192.168.10.10\учебный центр\Учеба\Журналы\XML'
WB_SHEET = 'с 09.01.2024'


def create_xml(prot: str, wb) -> str or bool:
    """
    Функция создания xml

    Создает файл <prot>.xml в указанной директории PATH
    """

    r = wb[WB_SHEET]

    # Создаем XML-элементы в соответствии с заданной схемой
    attr_qname = ET.QName("http://www.w3.org/2001/XMLSchema-instance", "noNamespaceSchemaLocation")

    registry_set = ET.Element('RegistrySet', {attr_qname: 'schema.xsd'},
                              nsmap={'xsi': 'http://www.w3.org/2001/XMLSchema-instance'})

    const_date = ''  # Дата проверки знаний
    const_n_prot = ''  # номер протокола проверки знаний

    # Перебор строк в excel
    for row in r.iter_rows(min_row=2):

        # Проверка условия на наличие имени в строке и проверка на загруженный протокол
        if row[3].value and not row[10].value and row[1].value == prot:

            # Специальные условия для получения констант
            if row[0].value:
                const_date = row[0].value
            if row[1].value:
                const_n_prot = row[1].value

            # Создание общего объекта
            registry_record = ET.SubElement(registry_set, 'RegistryRecord')

            # Создание работника
            worker = ET.SubElement(registry_record, 'Worker')

            # Разбитие фио на Ф И О
            fio = row[3].value.split()

            last_name = ET.SubElement(worker, 'LastName')
            last_name.text = fio[0]

            first_name = ET.SubElement(worker, 'FirstName')
            first_name.text = fio[1]

            middle_name = ET.SubElement(worker, 'MiddleName')

            try:
                match len(fio):

                    # Если ФИО состоит из 3 элементов
                    case 3:
                        middle_name.text = fio[2]

                    # Если ФИО состоит из 2 элементов
                    case 2:
                        middle_name.text = ''

                    # Проблемный кейс
                    case _:
                        return f'Проблемы с именем у ученика: {" ".join(fio)}.'

            except IndexError:
                return f'Проблемы с именем у ученика: {" ".join(fio)}.'

            # СНИЛС
            snils = ET.SubElement(worker, 'Snils')
            try:
                snils.text = row[5].value
            except KeyError:
                return f'Проблемы с СНИЛС: {" ".join(fio)}. Протокол {prot}'

            # Рабочее место
            position = ET.SubElement(worker, 'Position')
            try:
                position.text = row[4].value
            except KeyError:
                return f'Проблемы с РМ: {" ".join(fio)}. Протокол {prot}'

            # ИНН организации
            employer_inn = ET.SubElement(worker, 'EmployerInn')
            try:
                employer_inn.text = str(row[7].value)
            except KeyError:
                return f'Проблемы с ИНН организации: {" ".join(fio)}. Протокол {prot}'

            # Наименование организации
            employer_title = ET.SubElement(worker, 'EmployerTitle')

            try:
                employer_title.text = row[6].value
            except KeyError:
                return f'Проблемы с наименование организации: {" ".join(fio)}. Протокол {prot}'

            # Создание объекта организации
            organization = ET.SubElement(registry_record, 'Organization')
            inn = ET.SubElement(organization, 'Inn')
            inn.text = '3123356468'  # Наш ИНН

            title = ET.SubElement(organization, 'Title')
            title.text = 'ООО "АТМ"'  # Наше название

            if row[8].value == 'удовлетворительно':
                res_ob = 'true'
            else:
                res_ob = 'false'

            # Проверка правильности программы обучения
            try:
                prog_ob = str(row[9].value)
            except KeyError:
                return f'Учебной программы этого ученика нет в списке: {" ".join(fio)}.'

            # Создаем объект тест, указываем результат тестирования и номер программы обучения
            test = ET.SubElement(registry_record, 'Test', isPassed=res_ob, learnProgramId=prog_ob)

            # Создаем объект даты и заполняем
            date = ET.SubElement(test, 'Date')

            try:
                date.text = str(const_date).split()[0]
            except IndexError:
                return f'Дата в ячейке у этого ученика отсутствует: {" ".join(fio)}.'

            # Создаем объект протокола
            protocol_number = ET.SubElement(test, 'ProtocolNumber')
            try:
                protocol_number.text = const_n_prot
            except KeyError:
                return f'Проблемы с протоколом: {" ".join(fio)}. Протокол {prot}'

            # Указываем название программы обучения
            learn_program_title = ET.SubElement(test, 'LearnProgramTitle')

            try:
                learn_program_title.text = DI[str(row[9].value)]
            except KeyError:
                return f'Проблемы с программой обучения: {" ".join(fio)}. Протокол {prot}'

    # Создаем документ на основе элементов
    xml_doc = ET.ElementTree(registry_set)

    try:
        path = PATH_TO_XML + fr'\{prot.replace("/", " ")}.xml'
        # Сохраняем документ в файл
        xml_doc.write(path,
                      pretty_print=True, xml_declaration=True, encoding='utf-8')

        return True
    except OSError:
        return 'Ошибка доступа к папке для хранения xml'


def get_workbook():
    """
    Функция для получения книги по заданному пути в константах пути
    """

    try:
        return load_workbook(filename=PATH_TO_MAIN)
    except OSError:
        return 'Нет доступа к файлу, проверьте VPN или наличие файла в папке.'


def get_list_protocol(wb):
    """
    Функция для получения списка протоколов из книги
    """
    sheet = wb[WB_SHEET]

    protocol_list = sorted(
        set(
            (row[0].value, row[1].value) for row in sheet.iter_rows(min_row=2) if row[3].value and not row[10].value
        ),
        key=lambda x: x[0]
    )

    return (i[1] for i in protocol_list)


def data_update(f_name, wb_update):
    """
        Функция для загрузки списка протоколов из реестра
    """

    # Получаем книгу и открываем книгу
    wb = load_workbook(filename=f_name)
    sheet = wb.active

    if not wb_update:
        wb_update = load_workbook(filename=PATH_TO_MAIN)

    # Получаем итератор строк с незаполненными протоколами
    row_file = filter(lambda row_obj: not row_obj[10].value, wb_update['с 01.03.2023'].iter_rows(min_row=2))

    # Получаем итератор строк файла загрузки
    row_file_input = list(sheet.iter_rows(min_row=2))

    for row_main in row_file:

        # Находим и заполняем строку файла
        filter_obj = list(filter(
            lambda row_input: row_input[12].value == row_main[1].value and row_input[5].value == row_main[5].value,
            row_file_input))
        if filter_obj:
            row_main[10].value = filter_obj[0][0].value
        else:
            return f"Проблемы с загрузкой файла,проблема в строке СНИЛС: {row_main[5].value}"
    try:
        # Сохраняем файл
        wb_update.save(PATH_TO_MAIN)
        return "Сведения загружены успешно!"
    except OSError:
        return "Нет доступа к нашему excel"
