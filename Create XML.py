import lxml.etree as ET

import tkinter as tk
from tkinter import *
from tkinter import messagebox, filedialog

from openpyxl import load_workbook

di = {
    "1": "Оказание первой помощи пострадавшим",
    "2": "Использование (применение) средств индивидуальной защиты",
    "3": "Общие вопросы охраны труда и функционирования системы управления охраной труда",
    "4": "Безопасные методы и приемы выполнения работ при воздействии вредных и (или) опасных производственных факторов, источников опасности, идентифицированных в рамках специальной оценки условий труда и оценки профессиональных рисков",
    "6": "Безопасные методы и приемы выполнения земляных работ",
    "7": "Безопасные методы и приемы выполнения ремонтных, монтажных и демонтажных работ зданий и сооружений",
    "8": "Безопасные методы и приемы выполнения работ при размещении, монтаже, техническом обслуживании и ремонте технологического оборудования (включая технологическое оборудование)",
    "9": "Безопасные методы и приемы выполнения работ на высоте",
    "10": "Безопасные методы и приемы выполнения пожароопасных работ",
    "11": "Безопасные методы и приемы выполнения работ в ограниченных и замкнутых пространствах (ОЗП)",
    "12": "Безопасные методы и приемы выполнения строительных работ, в том числе: - окрасочные работы - электросварочные и газосварочные работы",
    "13": "Безопасные методы и приемы выполнения работ, связанных с опасностью воздействия сильнодействующих и ядовитых веществ",
    "14": "Безопасные методы и приемы выполнения газоопасных работ",
    "15": "Безопасные методы и приемы выполнения огневых работ",
    "16": "Безопасные методы и приемы выполнения работ, связанные с эксплуатацией подъемных сооружений",
    "17": "Безопасные методы и приемы выполнения работ, связанные с эксплуатацией тепловых энергоустановок",
    "18": "Безопасные методы и приемы выполнения работ в электроустановках",
    "19": "Безопасные методы и приемы выполнения работ, связанные с эксплуатацией сосудов, работающих под избыточным давлением",
    "20": "Безопасные методы и приемы обращения с животными",
    "21": "Безопасные методы и приемы при выполнении водолазных работ",
    "22": "Безопасные методы и приемы работ по поиску, идентификации, обезвреживанию и уничтожению взрывоопасных предметов",
    "23": "Безопасные методы и приемы работ в непосредственной близости от полотна или проезжей части эксплуатируемых автомобильных и железных дорог",
    "24": "Безопасные методы и приемы работ, на участках с патогенным заражением почвы",
    "25": "Безопасные методы и приемы работ по валке леса в особо опасных условиях",
    "26": "Безопасные методы и приемы работ по перемещению тяжеловесных и крупногабаритных грузов при отсутствии машин соответствующей грузоподъемности и разборке покосившихся и опасных (неправильно уложенных) штабелей круглых лесоматериалов",
    "27": "Безопасные методы и приемы работ с радиоактивными веществами и источниками ионизирующих излучений",
    "28": "Безопасные методы и приемы работ с ручным инструментом, в том числе с пиротехническим",
    "29": "Безопасные методы и приемы работ в театрах"
}


# Функция создания xml
def create_xml(prot, r):
    # Создаем XML-элементы в соответствии с заданной схемой
    attr_qname = ET.QName("http://www.w3.org/2001/XMLSchema-instance", "noNamespaceSchemaLocation")

    registry_set = ET.Element('RegistrySet', {attr_qname: 'schema.xsd'},
                              nsmap={'xsi': 'http://www.w3.org/2001/XMLSchema-instance'})

    const_date = ''  # Дата проверки знаний
    const_n_prot = ''  # номер протокола проверки знаний

    snils_list = []

    # Перебор строк в excel
    for row in r.iter_rows(min_row=2):

        # Проверка условия на наличие имени в строке и проверка на загруженный протокол
        if row[3].value and not row[10].value and row[1].value == prot:

            # Специальные условия для получения констант
            if row[0].value:
                const_date = row[0].value
            if row[1].value:
                const_n_prot = row[1].value

            # Создание общего объекта
            registry_record = ET.SubElement(registry_set, 'RegistryRecord')

            # Создание работника
            worker = ET.SubElement(registry_record, 'Worker')

            # Разбитие фио на Ф И О
            fio = row[3].value.split()

            last_name = ET.SubElement(worker, 'LastName')
            last_name.text = fio[0]

            first_name = ET.SubElement(worker, 'FirstName')
            first_name.text = fio[1]

            middle_name = ET.SubElement(worker, 'MiddleName')

            try:
                match len(fio):

                    # Если ФИО состоит из 3 элементов
                    case 3:
                        middle_name.text = fio[2]

                    # Если ФИО состоит из 2 элементов
                    case 2:
                        middle_name.text = ''

                    # Проблемный кейс
                    case _:
                        messagebox.showwarning("Предупреждение",
                                               f'Проблемы с именем у ученика: {" ".join(fio)}.')
                        return
            except IndexError:
                messagebox.showwarning("Предупреждение",
                                       f'Проблемы с именем у ученика: {" ".join(fio)}.')
                return

            # Проверка СНИЛС на повторение
            if row[5].value not in snils_list:
                snils_list.append(row[5].value)
            else:
                messagebox.showwarning("Предупреждение", f'Программа остановлена, СНИЛС повторяется: {row[5].value}.')
                return

            # СНИЛС
            snils = ET.SubElement(worker, 'Snils')
            snils.text = row[5].value

            # Рабочее место
            position = ET.SubElement(worker, 'Position')
            position.text = row[4].value

            # ИНН организации
            employer_inn = ET.SubElement(worker, 'EmployerInn')
            employer_inn.text = str(row[7].value)

            # Наименование организации
            employer_title = ET.SubElement(worker, 'EmployerTitle')
            employer_title.text = row[6].value

            # Создание объекта организации
            organization = ET.SubElement(registry_record, 'Organization')
            inn = ET.SubElement(organization, 'Inn')
            inn.text = '3123356468'  # Наш ИНН

            title = ET.SubElement(organization, 'Title')
            title.text = 'ООО "АТМ"'  # Наше название

            if row[8].value == 'удовлетворительно':
                res_ob = 'true'
            else:
                res_ob = 'false'

            # Проверка правильности программы обучения
            try:
                prog_ob = str(row[9].value)
            except KeyError:
                messagebox.showwarning("Предупреждение",
                                       f'Учебной программы этого ученика нет в списке: {" ".join(fio)}.')
                return

            # Создаем объект тест, указываем результат тестирования и номер программы обучения
            test = ET.SubElement(registry_record, 'Test', isPassed=res_ob, learnProgramId=prog_ob)

            # Создаем объект даты и заполняем
            date = ET.SubElement(test, 'Date')

            try:
                date.text = f'{str(const_date).split()[0]}T01:00:00'
            except IndexError:
                messagebox.showwarning("Предупреждение", f'Дата в ячейке у этого ученика отсутствует: {" ".join(fio)}.')
                return

            # Создаем объект протокола
            protocol_number = ET.SubElement(test, 'ProtocolNumber')
            protocol_number.text = const_n_prot

            # Указываем название программы обучения
            learn_program_title = ET.SubElement(test, 'LearnProgramTitle')
            learn_program_title.text = di[str(row[9].value)]

    # Создаем документ на основе элементов
    xml_doc = ET.ElementTree(registry_set)

    try:

        # Сохраняем документ в файл
        xml_doc.write(fr'\\192.168.10.10\учебный центр\Учеба\Журналы\XML\{prot.replace("/", " ")}.xml',
                      pretty_print=True, xml_declaration=True, encoding='utf-8')

        messagebox.showwarning("Сообщение", fr'Файл успешно сформирован {prot.replace("/", " ")}.xml')
    except OSError:

        messagebox.showwarning("Предупреждение", fr'Ошибка доступа к папке для хранения xml')
        return


def path_to_sheet():
    # Путь к файлу
    path = fr'\\192.168.10.10\учебный центр\Учеба\Журналы'

    try:

        # Загружаем книгу
        wb = load_workbook(filename=f'{path}\\Журнал регистрации удостоверений АТМ ДОТ 2023.xlsx')
    except OSError:
        messagebox.showwarning("Предупреждение", f'Книга Excel недоступна!!!')
        return

    # Выбираем необходимую страницу
    return wb['с 01.03.2023'], wb


# Создаем функцию для обработки нажатия на кнопку
def button_form_xml():
    sheet = path_to_sheet()[0]

    if sheet:

        # Вызываем функцию создания XML
        select = list(box.curselection())
        select.reverse()
        for i in select:
            create_xml(box.get(i), sheet)
    else:
        return


def button_list_protocol():
    # Выбираем лист
    sheet = path_to_sheet()[0]

    if sheet:

        # очистка списка протоколов
        box.delete(0, END)

        for row in sheet.iter_rows(min_row=2):

            # Проверка условия на наличие имени в строке и проверка на загруженный протокол
            if row[3].value and not row[10].value:
                if row[1].value not in box.get(0, END):
                    box.insert(END, row[1].value)
    else:
        return


def button_rev():
    # Выбор файла выгрузки
    file_path = filedialog.askopenfile(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
    wb = load_workbook(filename=file_path.name)
    sheet = wb.active

    # Получаем страницу и книгу
    sheet_dow, wb_dow = path_to_sheet()

    # Перебираем строки
    for row in sheet.iter_rows(min_row=2):
        for row_dow in sheet_dow.iter_rows(min_row=2):
            if not row_dow[10].value and row_dow[5].value == row[5].value and row_dow[1].value == row[11].value:
                row_dow[10].value = row[0].value

    try:

        # Сохраняем файл
        wb_dow.save(fr'\\192.168.10.10\учебный центр\Учеба\Журналы\Журнал регистрации удостоверений АТМ ДОТ 2023.xlsx')

        messagebox.showwarning("Сообщение", "Сведения загружены успешно!")

    except OSError:
        messagebox.showwarning("Предупреждение", f'Нет доступа к нашему excel!\nЗакройте файл!')
        return


if __name__ == '__main__':
    # Создаем окно
    window = tk.Tk()

    # Устанавливаем заголовок окна
    window.title("XML")

    # Устанавливаем размеры окна
    window.geometry("650x400")

    box = Listbox(selectmode=EXTENDED, width=50, height=50)
    box.pack(side=LEFT)

    scroll = Scrollbar(command=box.yview)
    scroll.pack(side=LEFT, fill=Y)

    box.config(yscrollcommand=scroll.set)

    # Создаем фрейм и размещаем его на форме
    f = tk.Frame(window)
    f.pack(side=LEFT, padx=10)

    # Создаем кнопку получения списка протоколов
    tk.Button(f, text="Получить список протоколов", command=button_list_protocol, fg='#ffffff', bg='#4CAF50', relief='flat',
              activebackground='#2E7D32', font=('Arial', 14)).pack(fill=X, pady=10)

    # Создаем кнопку формирования XML
    tk.Button(f, text="Сформировать XML", command=button_form_xml, fg='#ffffff', bg='#4CAF50', relief='flat',
              activebackground='#2E7D32', font=('Arial', 14)).pack(fill=X, pady=10)

    # Создаем кнопку обратной загрузки
    tk.Button(f, text="Обратная загрузка", command=button_rev, fg='#ffffff', bg='#4CAF50', relief='flat',
              activebackground='#2E7D32', font=('Arial', 14)).pack(fill=X, pady=10)

    # Запускаем главный цикл обработки событий
    window.mainloop()
